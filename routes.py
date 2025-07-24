from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from sqlalchemy import func, desc
from datetime import datetime, timedelta
import logging

from app import db
from models import User, Product, Category, Order, OrderItem, CartItem, Advertisement
from forms import LoginForm, RegisterForm, ProductForm, CategoryForm, AdvertisementForm, OrderForm



from flask import make_response, current_app
from datetime import datetime
import io
from reportlab.pdfgen import canvas  # لتوليد PDF
import openpyxl  # لتوليد Excel
import csv  # لتوليد CSV



# Main blueprint for public routes
main_bp = Blueprint('main', __name__)

# Admin blueprint for admin routes
admin_bp = Blueprint('admin', __name__)

# Helper function to check if user is admin
def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('ليس لديك صلاحية للوصول لهذه الصفحة', 'error')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function

# ============ MAIN ROUTES ============

@main_bp.route('/')
def index():
    # Get featured products
    featured_products = Product.query.filter_by(is_featured=True, is_active=True).limit(8).all()
    
    # Get active advertisements
    advertisements = Advertisement.query.filter_by(is_active=True).all()
    
    # Get categories
    categories = Category.query.filter_by(is_active=True).all()
    
    return render_template('index.html', 
                         featured_products=featured_products,
                         advertisements=advertisements,
                         categories=categories)

@main_bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.index'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter(
            (User.username == form.username.data) | 
            (User.email == form.username.data)
        ).first()
        
        if user and check_password_hash(user.password_hash, form.password.data):
            login_user(user)
            next_page = request.args.get('next')
            if user.is_admin:
                return redirect(next_page) if next_page else redirect(url_for('admin.dashboard'))
            return redirect(next_page) if next_page else redirect(url_for('main.index'))
        
        flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')
    
    return render_template('login.html', form=form)

@main_bp.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('main.index'))
    
    form = RegisterForm()
    if form.validate_on_submit():
        # Check if username exists
        if User.query.filter_by(username=form.username.data).first():
            flash('اسم المستخدم موجود مسبقاً', 'error')
            return render_template('register.html', form=form)
        
        # Check if email exists
        if User.query.filter_by(email=form.email.data).first():
            flash('البريد الإلكتروني موجود مسبقاً', 'error')
            return render_template('register.html', form=form)
        
        # Create new user
        user = User(
            username=form.username.data,
            email=form.email.data,
            password_hash=generate_password_hash(form.password.data),
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            phone=form.phone.data,
            address=form.address.data
        )
        
        db.session.add(user)
        db.session.commit()
        
        flash('تم تسجيل الحساب بنجاح', 'success')
        return redirect(url_for('main.login'))
    
    return render_template('register.html', form=form)

@main_bp.route('/logout')
@login_required
def logout():
    logout_user()
    flash('تم تسجيل الخروج بنجاح', 'success')
    return redirect(url_for('main.index'))

@main_bp.route('/products')
def products():
    page = request.args.get('page', 1, type=int)
    category_id = request.args.get('category', type=int)
    search = request.args.get('search', '')
    
    query = Product.query.filter_by(is_active=True)
    
    if category_id:
        query = query.filter_by(category_id=category_id)
    
    if search:
        query = query.filter(Product.name.contains(search))
    
    products = query.paginate(page=page, per_page=12, error_out=False)
    categories = Category.query.filter_by(is_active=True).all()
    
    return render_template('products.html', 
                         products=products,
                         categories=categories,
                         current_category=category_id,
                         search=search)

@main_bp.route('/product/<int:id>')
def product_detail(id):
    product = Product.query.get_or_404(id)
    related_products = Product.query.filter_by(
        category_id=product.category_id,
        is_active=True
    ).filter(Product.id != id).limit(4).all()
    
    return render_template('product_detail.html', 
                         product=product,
                         related_products=related_products)

@main_bp.route('/add_to_cart/<int:product_id>')
@login_required
def add_to_cart(product_id):
    product = Product.query.get_or_404(product_id)
    
    cart_item = CartItem.query.filter_by(
        user_id=current_user.id,
        product_id=product_id
    ).first()
    
    if cart_item:
        cart_item.quantity += 1
    else:
        cart_item = CartItem(
            user_id=current_user.id,
            product_id=product_id,
            quantity=1
        )
        db.session.add(cart_item)
    
    db.session.commit()
    flash('تم إضافة المنتج إلى السلة', 'success')
    return redirect(url_for('main.product_detail', id=product_id))

@main_bp.route('/cart')
@login_required
def cart():
    cart_items = CartItem.query.filter_by(user_id=current_user.id).all()
    total = sum(item.product.price * item.quantity for item in cart_items)
    
    return render_template('cart.html', 
                         cart_items=cart_items,
                         total=total)

@main_bp.route('/update_cart/<int:item_id>')
@login_required
def update_cart(item_id):
    quantity = request.args.get('quantity', 1, type=int)
    cart_item = CartItem.query.get_or_404(item_id)
    
    if cart_item.user_id != current_user.id:
        flash('غير مسموح', 'error')
        return redirect(url_for('main.cart'))
    
    if quantity <= 0:
        db.session.delete(cart_item)
    else:
        cart_item.quantity = quantity
    
    db.session.commit()
    return redirect(url_for('main.cart'))

@main_bp.route('/remove_from_cart/<int:item_id>')
@login_required
def remove_from_cart(item_id):
    cart_item = CartItem.query.get_or_404(item_id)
    
    if cart_item.user_id != current_user.id:
        flash('غير مسموح', 'error')
        return redirect(url_for('main.cart'))
    
    db.session.delete(cart_item)
    db.session.commit()
    flash('تم حذف المنتج من السلة', 'success')
    return redirect(url_for('main.cart'))

@main_bp.route('/checkout', methods=['GET', 'POST'])
@login_required
def checkout():
    cart_items = CartItem.query.filter_by(user_id=current_user.id).all()
    
    if not cart_items:
        flash('السلة فارغة', 'error')
        return redirect(url_for('main.cart'))
    
    form = OrderForm()
    if form.validate_on_submit():
        # Create order
        total = sum(item.product.price * item.quantity for item in cart_items)
        order = Order(
            user_id=current_user.id,
            total_amount=total,
            shipping_address=form.shipping_address.data,
            payment_method=form.payment_method.data,
            notes=form.notes.data
        )
        db.session.add(order)
        db.session.flush()
        
        # Create order items
        for cart_item in cart_items:
            order_item = OrderItem(
                order_id=order.id,
                product_id=cart_item.product_id,
                quantity=cart_item.quantity,
                price=cart_item.product.price
            )
            db.session.add(order_item)
        
        # Clear cart
        for cart_item in cart_items:
            db.session.delete(cart_item)
        
        db.session.commit()
        flash('تم تقديم الطلب بنجاح', 'success')
        return redirect(url_for('main.index'))
    
    total = sum(item.product.price * item.quantity for item in cart_items)
    return render_template('checkout.html', 
                         cart_items=cart_items,
                         total=total,
                         form=form)

# ============ ADMIN ROUTES ============

@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    # Statistics
    total_users = User.query.count()
    total_products = Product.query.count()
    total_orders = Order.query.count()
    total_revenue = db.session.query(func.sum(Order.total_amount)).scalar() or 0
    
    # Recent orders
    recent_orders = Order.query.order_by(desc(Order.created_at)).limit(10).all()
    
    # Monthly sales data for chart
    monthly_sales = db.session.query(
        func.date_trunc('month', Order.created_at).label('month'),
        func.sum(Order.total_amount).label('total')
    ).group_by('month').order_by('month').all()
    
    return render_template('admin/dashboard.html',
                         total_users=total_users,
                         total_products=total_products,
                         total_orders=total_orders,
                         total_revenue=total_revenue,
                         recent_orders=recent_orders,
                         monthly_sales=monthly_sales)

@admin_bp.route('/products')
@login_required
@admin_required
def products():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    
    query = Product.query
    if search:
        query = query.filter(Product.name.contains(search))
    
    products = query.paginate(page=page, per_page=20, error_out=False)
    
    return render_template('admin/products.html', 
                         products=products,
                         search=search)

@admin_bp.route('/products/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_product():
    form = ProductForm()
    form.category_id.choices = [(c.id, c.name) for c in Category.query.filter_by(is_active=True).all()]
    
    if form.validate_on_submit():
        product = Product(
            name=form.name.data,
            description=form.description.data,
            price=form.price.data,
            stock_quantity=form.stock_quantity.data,
            image_url=form.image_url.data,
            category_id=form.category_id.data,
            is_active=form.is_active.data,
            is_featured=form.is_featured.data
        )
        db.session.add(product)
        db.session.commit()
        flash('تم إضافة المنتج بنجاح', 'success')
        return redirect(url_for('admin.products'))
    
    return render_template('admin/add_product.html', form=form)

@admin_bp.route('/products/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_product(id):
    product = Product.query.get_or_404(id)
    form = ProductForm(obj=product)
    form.category_id.choices = [(c.id, c.name) for c in Category.query.filter_by(is_active=True).all()]
    
    if form.validate_on_submit():
        form.populate_obj(product)
        db.session.commit()
        flash('تم تحديث المنتج بنجاح', 'success')
        return redirect(url_for('admin.products'))
    
    return render_template('admin/edit_product.html', form=form, product=product)

@admin_bp.route('/products/delete/<int:id>')
@login_required
@admin_required
def delete_product(id):
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    flash('تم حذف المنتج بنجاح', 'success')
    return redirect(url_for('admin.products'))

@admin_bp.route('/orders')
@login_required
@admin_required
def orders():
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    
    query = Order.query
    if status:
        query = query.filter_by(status=status)
    
    orders = query.order_by(desc(Order.created_at)).paginate(page=page, per_page=20, error_out=False)
    
    return render_template('admin/orders.html', 
                         orders=orders,
                         current_status=status)

@admin_bp.route('/orders/update_status/<int:id>')
@login_required
@admin_required
def update_order_status(id):
    status = request.args.get('status')
    order = Order.query.get_or_404(id)
    order.status = status
    db.session.commit()
    flash('تم تحديث حالة الطلب بنجاح', 'success')
    return redirect(url_for('admin.orders'))

@admin_bp.route('/users')
@login_required
@admin_required
def users():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    
    query = User.query
    if search:
        query = query.filter(User.username.contains(search) | User.email.contains(search))
    
    users = query.paginate(page=page, per_page=20, error_out=False)
    
    return render_template('admin/users.html', 
                         users=users,
                         search=search)

@admin_bp.route('/advertisements')
@login_required
@admin_required
def advertisements():
    page = request.args.get('page', 1, type=int)
    advertisements = Advertisement.query.paginate(page=page, per_page=20, error_out=False)
    
    return render_template('admin/advertisements.html', 
                         advertisements=advertisements)

@admin_bp.route('/advertisements/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_advertisement():
    form = AdvertisementForm()
    
    if form.validate_on_submit():
        advertisement = Advertisement(
            title=form.title.data,
            description=form.description.data,
            image_url=form.image_url.data,
            link_url=form.link_url.data,
            position=form.position.data,
            is_active=form.is_active.data,
            start_date=form.start_date.data,
            end_date=form.end_date.data
        )
        db.session.add(advertisement)
        db.session.commit()
        flash('تم إضافة الإعلان بنجاح', 'success')
        return redirect(url_for('admin.advertisements'))
    
    return render_template('admin/add_advertisement.html', form=form)




admin_bp.route('/reports')
@login_required
@admin_required
def reports():
    # Sales report
    total_sales = db.session.query(func.sum(Order.total_amount)).scalar() or 0
    orders_count = Order.query.count()
    
    # Top products
    top_products = db.session.query(
        Product.name,
        func.sum(OrderItem.quantity).label('total_sold')
    ).join(OrderItem).group_by(Product.name).order_by(desc('total_sold')).limit(10).all()
    
    # Monthly sales
    monthly_sales = db.session.query(
        func.date_trunc('month', Order.created_at).label('month'),
        func.sum(Order.total_amount).label('total')
    ).group_by('month').order_by('month').all()
    
    return render_template('admin/reports.html',
                         total_sales=total_sales,
                         orders_count=orders_count,
                         top_products=top_products,
                         monthly_sales=monthly_sales)


@admin_bp.route('/export_report/<format>')
@login_required
@admin_required
def export_report(format):
    """تصدير التقارير بصيغ مختلفة"""
    try:
        # جمع البيانات
        total_sales = db.session.query(func.sum(Order.total_amount)).scalar() or 0
        orders_count = Order.query.count()
        
        # Top products
        top_products = db.session.query(
            Product.name,
            func.sum(OrderItem.quantity).label('total_sold')
        ).join(OrderItem).group_by(Product.name).order_by(desc('total_sold')).limit(10).all()
        
        # Monthly sales
        monthly_sales = db.session.query(
            func.date_trunc('month', Order.created_at).label('month'),
            func.sum(Order.total_amount).label('total')
        ).group_by('month').order_by('month').all()
        
        # All orders for detailed report
        orders = db.session.query(
            Order.id,
            Order.created_at,
            Order.total_amount,
            Order.status,
            User.username
        ).join(User).order_by(Order.created_at.desc()).all()
        
        if format.lower() == 'csv':
            return export_csv_report(orders, monthly_sales, top_products)
        elif format.lower() == 'excel':
            return export_excel_report(orders, monthly_sales, top_products)
        elif format.lower() == 'pdf':
            return export_pdf_report(orders, monthly_sales, top_products, total_sales, orders_count)
        else:
            flash('صيغة التصدير غير مدعومة', 'error')
            return redirect(url_for('admin.reports'))
            
    except Exception as e:
        flash(f'خطأ في تصدير التقرير: {str(e)}', 'error')
        return redirect(url_for('admin.reports'))

def export_csv_report(orders, monthly_sales, top_products):
    """تصدير التقرير بصيغة CSV"""
    import csv
    import io
    from flask import make_response
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # كتابة بيانات الطلبات
    writer.writerow(['رقم الطلب', 'التاريخ', 'المبلغ', 'الحالة', 'اسم المستخدم'])
    for order in orders:
        writer.writerow([
            order.id,
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            float(order.total_amount),
            order.status,
            order.username
        ])
    
    # إضافة فاصل
    writer.writerow([])
    writer.writerow(['المبيعات الشهرية'])
    writer.writerow(['الشهر', 'إجمالي المبيعات'])
    for sale in monthly_sales:
        writer.writerow([
            sale.month.strftime('%Y-%m'),
            float(sale.total or 0)
        ])
    
    # إضافة أفضل المنتجات
    writer.writerow([])
    writer.writerow(['أفضل المنتجات'])
    writer.writerow(['اسم المنتج', 'الكمية المباعة'])
    for product in top_products:
        writer.writerow([product.name, product.total_sold])
    
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{datetime.now().strftime("%Y%m%d")}.csv'
    
    return response

def export_excel_report(orders, monthly_sales, top_products):
    """تصدير التقرير بصيغة Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        import io
        from flask import make_response
        
        # إنشاء workbook جديد
        wb = openpyxl.Workbook()
        
        # ورقة الطلبات
        ws_orders = wb.active
        ws_orders.title = "الطلبات"
        
        # عناوين الأعمدة
        headers = ['رقم الطلب', 'التاريخ', 'المبلغ', 'الحالة', 'اسم المستخدم']
        for col, header in enumerate(headers, 1):
            cell = ws_orders.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # بيانات الطلبات
        for row, order in enumerate(orders, 2):
            ws_orders.cell(row=row, column=1, value=order.id)
            ws_orders.cell(row=row, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws_orders.cell(row=row, column=3, value=float(order.total_amount))
            ws_orders.cell(row=row, column=4, value=order.status)
            ws_orders.cell(row=row, column=5, value=order.username)
        
        # ورقة المبيعات الشهرية
        ws_monthly = wb.create_sheet("المبيعات الشهرية")
        ws_monthly.cell(row=1, column=1, value="الشهر").font = Font(bold=True)
        ws_monthly.cell(row=1, column=2, value="إجمالي المبيعات").font = Font(bold=True)
        
        for row, sale in enumerate(monthly_sales, 2):
            ws_monthly.cell(row=row, column=1, value=sale.month.strftime('%Y-%m'))
            ws_monthly.cell(row=row, column=2, value=float(sale.total or 0))
        
        # ورقة أفضل المنتجات
        ws_products = wb.create_sheet("أفضل المنتجات")
        ws_products.cell(row=1, column=1, value="اسم المنتج").font = Font(bold=True)
        ws_products.cell(row=1, column=2, value="الكمية المباعة").font = Font(bold=True)
        
        for row, product in enumerate(top_products, 2):
            ws_products.cell(row=row, column=1, value=product.name)
            ws_products.cell(row=row, column=2, value=product.total_sold)
        
        # حفظ في memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        return response
        
    except ImportError:
        flash('مكتبة openpyxl غير مثبتة. يرجى تثبيتها لتصدير ملفات Excel', 'error')
        return redirect(url_for('admin.reports'))

def export_pdf_report(orders, monthly_sales, top_products, total_sales, orders_count):
    """تصدير التقرير بصيغة PDF"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import io
        from flask import make_response
        
        # إنشاء buffer
        buffer = io.BytesIO()
        
        # إنشاء PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # الأنماط
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # العنوان
        title = Paragraph("تقرير المبيعات", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # الملخص
        summary_data = [
            ['إجمالي المبيعات', f'${total_sales:.2f}'],
            ['عدد الطلبات', str(orders_count)],
            ['متوسط قيمة الطلب', f'${(total_sales/orders_count if orders_count > 0 else 0):.2f}']
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # أفضل المنتجات
        if top_products:
            story.append(Paragraph("أفضل المنتجات مبيعاً", styles['Heading2']))
            products_data = [['اسم المنتج', 'الكمية المباعة']]
            for product in top_products[:5]:  # أول 5 منتجات
                products_data.append([product.name, str(product.total_sold)])
            
            products_table = Table(products_data)
            products_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(products_table)
        
        # بناء PDF
        doc.build(story)
        
        # إرجاع الاستجابة
        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{datetime.now().strftime("%Y%m%d")}.pdf'
        
        return response
        
    except ImportError:
        flash('مكتبة reportlab غير مثبتة. يرجى تثبيتها لتصدير ملفات PDF', 'error')
        return redirect(url_for('admin.reports'))

